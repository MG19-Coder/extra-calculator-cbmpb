import json
from pathlib import Path

import openpyxl

workbook_path = Path("Valor do Extra 2024 + Conversor.xlsx")
workbook = openpyxl.load_workbook(workbook_path, data_only=True)
sheet = workbook["1"]

normal_rows = range(17, 22)
majorado_rows = range(24, 29)
normal_by_grad = {sheet.cell(row, 1).value: sheet.cell(row, 2).value for row in normal_rows}
normal_hour_by_grad = {sheet.cell(row, 8).value: sheet.cell(row, 9).value for row in normal_rows}
majorado_by_grad = {sheet.cell(row, 1).value: sheet.cell(row, 2).value for row in majorado_rows}
majorado_hour_by_grad = {sheet.cell(row, 8).value: sheet.cell(row, 9).value for row in majorado_rows}

class_values = {
    "CFSD": sheet.cell(21, 21).value / sheet.cell(21, 20).value,
    "CFS": sheet.cell(22, 21).value / sheet.cell(22, 20).value,
    "CFO": sheet.cell(23, 21).value / sheet.cell(23, 20).value,
}

rows = []
for graduacao in normal_by_grad:
    rows.append({
        "graduacao": graduacao,
        "valorHoraExtraNormal": normal_hour_by_grad[graduacao],
        "valorExtraNormal12h": normal_by_grad[graduacao] / 2,
        "valorExtraNormal24h": normal_by_grad[graduacao],
        "valorHoraExtraMajorado": majorado_hour_by_grad[graduacao],
        "valorExtraMajorado12h": majorado_by_grad[graduacao] / 2,
        "valorExtraMajorado24h": majorado_by_grad[graduacao],
        "valorHoraAulaCFSD": class_values["CFSD"],
        "valorHoraAulaCFS": class_values["CFS"],
        "valorHoraAulaCFO": class_values["CFO"],
    })

print(json.dumps(rows, ensure_ascii=False, indent=2))
